import openpyxl
from googletrans import LANGUAGES


class tag:
    """Holds data about a tag"""

    def __init__(
        self,
        typ: str,
        content: str,
    ):
        self.typ = typ
        self.content = content


# ? provide this with the correct language code at the end of the base file name
baseFileName = "exspiron2xi"

#! make sure to set this to make the file structure work
lang = "de"


fileFound = False

try:
    # open the file
    with open(
        "{}_{}.ts".format(baseFileName, lang), "r", encoding="utf-8"
    ) as inputdata:

        # setup the workbook and it's variables
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Translations"

        # setup some basic variables
        extractedTags = []
        data = open(
            "{}_{}.ts".format(baseFileName, lang), "r", encoding="utf-8"
        ).readlines()

        # iterate through the lines of the file
        for lineIndex, line in enumerate(inputdata):

            # find the specific tags
            if (
                "<source" in line
                or "<translation" in line
                or "<translatorcomment" in line
            ):

                # set these up outside class definition for use in class definition
                contentStart = line.find(">") + 1
                contentEnd = line.find("<", contentStart)

                # check to see if we even need to do lookahead
                if (
                    "</source" not in line
                    and "</translation" not in line
                    and "</translatorcomment" not in line
                ):

                    # if we do, setup a few lines of lookahead
                    lookahead = data[lineIndex : lineIndex + 4]

                    # loop through the lookahead
                    for jndex, lineahead in enumerate(lookahead):

                        # find the end tag
                        if (
                            "</source" in lineahead
                            or "</translation" in lineahead
                            or "</translatorcomment" in lineahead
                        ):

                            # merge the range of lines and cut out the content
                            contentRange = "".join(
                                data[lineIndex : jndex + lineIndex + 1]
                            )
                            contentStart = contentRange.find(">") + 1
                            contentEnd = contentRange.find("<", contentStart)
                            contentRange = contentRange[contentStart:contentEnd]
                            break
                else:
                    contentRange = line[contentStart:contentEnd]

                # build the list of tags
                extractedTags.append(
                    tag(
                        line[line.find("<") + 1 : line.find(">")],
                        contentRange,
                    )
                )

        # setup a list to hold the indices of the tags that don't have a comment
        indicies = []

        # loop through all the tags
        for index, tg in enumerate(extractedTags):

            # prevent an index out of range error
            if index < len(extractedTags) - 1:
                nexttag = extractedTags[index + 1]
            else:
                nexttag = None

            # check if we're at a source tag and there is a next tag
            if tg.typ == "source" and nexttag != None:

                # if we're missing a comment,
                if nexttag.typ != "translatorcomment":

                    # add the index to the list of indices
                    indicies.append(index + 1)

        # loop through all the indices, with and index into that list for correction
        for index, position in enumerate(indicies):

            # insert a blank comment
            extractedTags.insert(position, tag("translatorcomment", ""))

            # add 1 to the rest of the indices down the list to allow for correct indexing further down
            for jndex, number in enumerate(indicies[index:]):
                indicies[jndex + index] = number + 1

        # reformat the extractedTags list for easier use later down
        extractedTags = [
            [tag for tag in extractedTags if tag.typ == "source"],
            [tag for tag in extractedTags if "translation" in tag.typ],
            [tag for tag in extractedTags if tag.typ == "translatorcomment"],
        ]

        # fix the all te ascii escape characters
        for tg in extractedTags[0] + extractedTags[2]:
            tg.content = (
                tg.content.replace("&quot;", '"')
                .replace("&apos;", "'")
                .replace("&gt;", "<")
                .replace("&lt;", ">")
            )

        # load the correct amount of cells into memory
        for x in range(1, 4):
            for y in range(1, len(extractedTags[0]) + 1):
                ws.cell(row=x, column=y)

        # make the title cells
        ws["A1"] = "Source"
        ws["C1"] = "Comments"

        # write all the sources to the first column of the worksheet
        for index in range(len(extractedTags[0])):
            ws["A{}".format(index + 3)] = extractedTags[0][index].content

        # write all the comments to the third row of the worksheet
        for index in range(len(extractedTags[2])):
            ws["C{}".format(index + 3)] = extractedTags[2][index].content

        # save the file
        wb.save("{}_{}_edited.xlsx".format(baseFileName, lang))


except Exception as xcpt:
    # oopsies!
    print(xcpt)
    input()
