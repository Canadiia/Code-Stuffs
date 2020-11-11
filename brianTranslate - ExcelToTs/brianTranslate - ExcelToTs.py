#!/usr/bin/python3

"""This script take translations from an excel file and put them into a Qt .ts file.
See the `print_usage_message()` function for how to invoke this script.

For the format that the Excel file needs to be in, see the `get_data` function.
"""

import re
import sys
from lxml import etree as xml
from typing import List, NamedTuple, Optional

import openpyxl
from googletrans import Translator
from openpyxl.workbook.workbook import Workbook


class Arguments(NamedTuple):
    """holds all the data extracted from the command line arguments passed in"""

    excel_filepath: str
    ts_filepath: str
    verification: bool


class Translation:
    """This holds the information gathered from the Excel file"""

    def __init__(
        self, english_source: str, translation: str, translator_comment: str
    ) -> None:
        self.english_source = english_source
        self.translation = translation
        self.translator_comment = translator_comment


def get_args(raw_args: List[str]) -> Optional[Arguments]:
    """This grabs the command line arguments parsed in alongside the file

    Args:
        raw_args (List[str]): the raw args sent in when running the file

    Returns:
        Optional[Arguments]: returns an `Argument` class if there are arguments
        or `None` if there was an issue
    """

    # see if we gave any command line arguments
    argc = len(sys.argv)
    if argc < 3:
        print("Not enough arguments.")
        print_usage_message()
        return None

    # Validate them
    if not raw_args[1].lower().endswith(".xlsx"):
        print("First argument must be to an Excel file.")
        print_usage_message()
        return None

    if not raw_args[2].lower().endswith(".ts"):
        print("Second argument must be to a Qt .ts XML file.")
        print_usage_message()
        return None

    # handle if we just don't give a 3rd argument
    if argc == 3:
        return Arguments(
            excel_filepath=raw_args[1],
            ts_filepath=raw_args[2],
            verification=False,
        )
    else:
        return Arguments(
            excel_filepath=raw_args[1],
            ts_filepath=raw_args[2],
            verification=raw_args[3].lower() == "-v",
        )


def print_usage_message() -> None:
    """this prints out the usage of the file"""
    print("Usage:")
    print("  <filename>.py <translation_source>.xlsx <destination>.ts [-v]")
    print("  the -v is optional and enable the output of a verification file")


def get_data(filepath: str) -> List[Translation]:
    """this will take in a filepath to an Excel file and grab the data from it
    and put it into a Translation class list

    Args:
        filepath (str): the filepath to the Excel file

    Raises:
        Exception: this is raised if there data in one of the cells but not both

    Returns:
        List[Translation]: this will hold all of the translations gathered from the Excel file
    """

    # load the Excel file located at the filepath provided in read only mode
    # and get the current active Sheet in the Excel file
    input_data = openpyxl.load_workbook(filename=filepath, read_only=True)
    active_sheet = input_data.active

    # create the empty data list
    data = []

    # load the rows
    max_rows = active_sheet.max_row + 1
    rows = active_sheet["A3:D%i" % max_rows]

    # loop through the rows
    for english_cell, translation_cell, _, comment_cell in rows:
        english_source = english_cell.value
        translation = translation_cell.value
        translator_comment = comment_cell.value

        if not english_source and not translation:
            # if the cells are empty, we skip it
            continue

        elif english_source and translation:
            # if we have data in both cells, add a `Translation` class to the
            #   data list
            data.append(
                Translation(
                    english_source,
                    translation,
                    translator_comment,
                )
            )

        else:
            # This is bad, one of them is None and the other has a value, ERROR
            raise Exception(
                (
                    "Mismatching English and Translation; one cel is empty",
                    english_source,
                    translation,
                )
            )

    # close the Excel file as loading it in read only mode requires closing
    input_data.close()
    return data


def get_translations(translation_class_list: List[Translation]) -> List[str]:
    """this takes in a list of Translation classes and returns a list
    containing just the "translation" variable from them

    Args:
        translation_class_list (list): A list of Translation classes

    Returns:
        list: A list containing just the "translation" variables
        from the original list
    """
    return [i.translation for i in translation_class_list]


def html_prep(input_string: str) -> str:
    """The Excel file uses characters that aren't allowed in the .ts file.
    This will take in an input string and replace
    all the nonvalid characters with valid ones

    Args:
        input_string (str): The string to be cleansed of it's sins

    Returns:
        string: the cleaned string
    """
    return (
        input_string.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
        .replace("\xa0", "&#160;")
    )


def cleanup_translation(input_string: str) -> str:
    """This takes in a string and cleans out all the escape characters
    for use in more accurate translations

    Args:
        input_string (str): the string to be cleaned

    Returns:
        string: the cleaned string
    """
    # search for any html tags
    html_tag = re.findall("<[^>]*>", input_string)

    # if there's any, loop through them and convert them into a space
    if html_tag:
        for sequence in html_tag:
            input_string = input_string.replace(sequence, " ")

    # fix the common escape sequences
    input_string = input_string.replace("\n", " ").replace("\xa0", " ")

    # get rid of the "%" references
    for i in range(1, 11):
        input_string = input_string.replace("%{}".format(i), " ")

    return input_string


def translate_translations(translations_list: List[str]) -> List[str]:
    """this takes in a list of strings to be translated and translates them

    Args:
        translations_list (list): a list of strings to be translated

    Returns:
        list: a list of the translated strings
    """
    translations = Translator().translate(translations_list, dest="english")
    return [translation.text for translation in translations]


def verification_file(
    translation_class_list: List[Translation], translations: List[str], language: str
) -> None:
    """this creates an Excel file for organization of the verification data

    Args:
        translation_class_list (list): the list of the Translation classes
        language (str): the language that we translated from
    """

    # create an empty Excel file, name the sheet and select the current sheet
    output_workbook = Workbook()
    output_workbook.title = "Verification Data"
    output_worksheet = output_workbook.active

    # load the correct amount of rows and columns
    for x in range(1, 3):
        for y in range(1, len(translation_class_list) + 1):
            output_worksheet.cell(row=x, column=y)

    # label the columns
    output_worksheet["A1"] = "Source"
    output_worksheet["B1"] = "Translated Translation"

    # put the data in the Excel file with the original source text in the
    #   A column and the translated translation in the B column
    max_rows = len(translation_class_list) + 1
    rows = output_worksheet["A2:B%i" % max_rows]
    for index, (english_source, translated_translation) in enumerate(rows):
        english_source.value = translation_class_list[index].english_source
        translated_translation.value = translations[index]

    # save the file
    output_workbook.save("Verification_Data_{}.xlsx".format(language))


def write_to_ts_file(filepath: str, translation_class_list: List[Translation]) -> None:
    """this takes the input .ts filepath and puts the Excell data
    into its coresponding location

    Args:
        filepath (str): the filepath to the .ts file
        translation_class_list (list): thi list of Translation classes
    """

    # Load the .ts file and make a variable for it's root
    ts_xml = xml.parse(filepath)
    all_source_tags = ts_xml.findall(".//source")

    # Go through each translation entry
    for translation in translation_class_list:
        # Find the <source> tags that have the matching english
        matching_tags = filter(
            lambda x: (x.text == translation.english_source), all_source_tags
        )

        # For each of the matching ones, edit the <translation> tag to use the translation
        for tag in matching_tags:
            message_tag = tag.getparent()
            message_tag.find("translation").text = translation.translation

    # save the file
    ts_xml.write(filepath, pretty_print=True, xml_declaration=True, encoding="utf-8")


def main() -> None:
    """this is the main function that handles
    the ordering of the other functions"""

    # we first get the file path of the Excel file, .ts file, and
    #   wether or not we want a verification file
    args = get_args(sys.argv)
    if not args:
        print("Error; arguments incorrect. Exiting.", file=sys.stderr)
        sys.exit(1)

    # next, we get a list of Translation classes containing
    #   the translations in the excel file.
    # we also get the current language from the end of the file
    #   for use in naming a verification file if we're making one
    translation_records = get_data(args.excel_filepath)

    # if it does, first we see if the user wants a verification file
    if args.verification:

        # if they do, First we get all the translations from
        #   the list of Translation classes
        translations = get_translations(translation_records)

        # next we strip out all the escape characters to make the process
        #   of translating more accurate
        translations = [
            cleanup_translation(translation) for translation in translations
        ]

        # now we translate the translations back to english
        translations = translate_translations(translations)

        # Finally we write the data to a verification Excel file
        language = args.excel_filepath[-7:-5]
        verification_file(translation_records, translations, language)

    # next we need to make sure that the text fits within the constraints
    #   of the .ts file
    # we do this by replacing the symbols in the text to
    #   HTML compliant ones
    for translation in translation_records:
        translation.translation = html_prep(translation.translation)

    # finally we write the data from the excel file to the .ts file
    write_to_ts_file(args.ts_filepath, translation_records)


if __name__ == "__main__":
    main()
