import googletrans
import openpyxl
from googletrans import Translator

# Translate the first column in the excel file into whatever language is needed and put it into the second column

# take in a list of sentences and translate them
def translate(sentences: list, desiredLang: str):
    translator = Translator()
    translations = translator.translate(sentences, dest=desiredLang)
    return [translation.text for translation in translations]


# get the desired language
deciding = True
while deciding:
    desiredLang = input(
        'what is the desired lang?\nenter "1" if you need to see all the language codes\n'
    )
    if desiredLang.isdigit():
        if int(desiredLang) == 1:
            for lang in googletrans.LANGCODES:
                print(lang)
        else:
            print("Invalid Input")
    else:
        desiredLang = desiredLang.lower()
        if desiredLang in googletrans.LANGCODES:
            deciding = False
        else:
            print("invalid language code")

# load the data
try:
    ws = openpyxl.load_workbook("1Xi_Translations_Pediatrics.xlsx")
except:
    print("Can't locate input file\nPress enter to close\n")
    input()

# clean the data

inputData = ws["Sheet2"]["A"]
formattedData = [cell.internal_value for cell in inputData[2:]]

# translate the data

formattedData = translate(formattedData, desiredLang)

# put the data into the translations column

outputColumn = ws["Sheet2"]["B"]
ws["Sheet2"]["B2"] = desiredLang.capitalize()
for index in range(len(outputColumn[2:])):
    ws["Sheet2"]["B{}".format(index + 3)] = formattedData[index]

# save and quit

ws.save("1Xi_Translations_Pediatrics_translated.xlsx")
ws.close()